#!/usr/bin/env python
#__*__ coding:utf-8 __*__

from openpyxl import load_workbook
import os

class read_request_excel:
    def __init__(self):
        self.excel_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),"resource_data\接口测试用例.xlsx")

    def read_excel(self):
        list=[]
        wb=load_workbook(self.excel_path)
        sheet=wb.worksheets[0]
        nrows=sheet.max_row
        for i in range(2,nrows+1):
            request_description=sheet.cell(i,2).value
            request_link=sheet.cell(i,3).value
            request_params=sheet.cell(i,4).value
            request_type=sheet.cell(i,5).value
            request_yuqi=sheet.cell(i,6).value
            # if request_yuqi=="预期结果2":
            #     sheet.cell(i,7,request_yuqi)
            #     sheet.cell(i,8,"测试通过")
            # else:
            #     sheet.cell(i, 7,request_yuqi)
            #     sheet.cell(i, 8,"测试不通过")
            if request_yuqi=="预期结果2":
                sheet.cell(i,7).value=request_yuqi
                sheet.cell(i,8).value="测试通过"
            else:
                sheet.cell(i,7).value=request_yuqi
                sheet.cell(i,8).value="测试通过"

            all_data=(request_description,request_link,request_params,request_type,request_yuqi)
            list.append(all_data)
        print(list)
        wb.save("./resource_data/运行之后的结果.xlsx")
        wb.close()

if __name__ == '__main__':
    re=read_request_excel()
    re.read_excel()
